import pickle
import numpy as np
from scipy import interpolate
from openpyxl import load_workbook
from numpy.random import multivariate_normal as mvnrnd
from scipy.stats import wishart
from scipy.stats import invwishart
from numpy.linalg import inv as inv


def aggregate(a, n=3):
    cumsum = np.cumsum(a, dtype=float)
    ret = []
    for i in range(-1, len(a) - n, n):
        low_index = i if i >= 0 else 0
        ret.append(cumsum[low_index + n] - cumsum[low_index])
    return ret


def loadWithoutZF(
    file_path,
    sheet_name="Sheet1",
    aggre_delta=1,
    verbose=True
):
    wb = load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    data = []
    data_head = []
    for row_values in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True
    ):
        data_head.append(row_values[0])
        data.append(row_values[1:])
    wb.close()
    if aggre_delta != 1:
        data = [aggregate(row, aggre_delta) for row in data]
    data = np.array(data).T
    return data, data_head


def kr_prod(a, b):
    return np.einsum('ir, jr -> ijr', a, b).reshape(a.shape[0] * b.shape[0], -1)


def cov_mat(mat):
    dim1, dim2 = mat.shape
    new_mat = np.zeros((dim2, dim2))
    mat_bar = np.mean(mat, axis = 0)
    for i in range(dim1):
        new_mat += np.einsum('i, j -> ij', mat[i, :] - mat_bar, mat[i, :] - mat_bar)
    return new_mat

def mat2ten(mat, tensor_size, mode):
    index = list()
    index.append(mode)
    for i in range(tensor_size.shape[0]):
        if i != mode:
            index.append(i)
    return np.moveaxis(np.reshape(mat, list(tensor_size[index]), order = 'F'), 0, mode)


def mnrnd(M, U, V):
    """
    Generate matrix normal distributed random matrix.
    M is a m-by-n matrix, U is a m-by-m matrix, and V is a n-by-n matrix.
    """
    dim1, dim2 = M.shape
    X0 = np.random.rand(dim1, dim2)
    P = np.linalg.cholesky(U)
    Q = np.linalg.cholesky(V)
    return M + np.matmul(np.matmul(P, X0), Q.T)


def ten2mat(tensor, mode):
    return np.reshape(np.moveaxis(tensor, mode, 0), (tensor.shape[mode], -1), order = 'F')


def BPMF(dense_mat, sparse_mat, init, rank, maxiter1, maxiter2):
    """Bayesian Probabilistic Matrix Factorization, BPMF."""
    W = init["W"]
    X = init["X"]
    dim1, dim2 = sparse_mat.shape
    dim = np.array([dim1, dim2])
    pos = np.where((dense_mat != 0) & (sparse_mat == 0))
    position = np.where(sparse_mat != 0)
    binary_mat = np.zeros((dim1, dim2))
    binary_mat[position] = 1
        
    beta0 = 1
    nu0 = rank
    mu0 = np.zeros((rank))
    W0 = np.eye(rank)
    tau = 1
    alpha = 1e-6
    beta = 1e-6
    
    W_plus = np.zeros((dim1, rank))
    X_plus = np.zeros((dim2, rank))
    mat_hat_plus = np.zeros((dim1, dim2))
    for iters in range(maxiter1):
#         import pdb; pdb.set_trace()
        for order in range(2):
            if order == 0:
                mat = W.copy()
            elif order == 1:
                mat = X.copy()
            mat_bar = np.mean(mat, axis = 0)
            var_mu_hyper = (dim[order] * mat_bar + beta0 * mu0)/(dim[order] + beta0)
            var_W_hyper = inv(inv(W0) + cov_mat(mat) + dim[order] * beta0/(dim[order] + beta0)
                             * np.outer(mat_bar - mu0, mat_bar - mu0))
            var_Lambda_hyper = wishart(df = dim[order] + nu0, scale = var_W_hyper, seed = None).rvs()
            var_mu_hyper = mvnrnd(var_mu_hyper, inv((dim[order] + beta0) * var_Lambda_hyper))
            
            if order == 0:
                var1 = X.T
                mat0 = np.matmul(var1, sparse_mat.T)
            elif order == 1:
                var1 = W.T
                mat0 = np.matmul(var1, sparse_mat)
            var2 = kr_prod(var1, var1)
            if order == 0:
                mat1 = np.matmul(var2, binary_mat.T)
            elif order == 1:
                mat1 = np.matmul(var2, binary_mat)
            var3 = tau * mat1.reshape(rank, rank, dim[order]) + np.dstack([var_Lambda_hyper] * dim[order])
            var4 = tau * mat0 + np.dstack([np.matmul(var_Lambda_hyper, var_mu_hyper)] * dim[order])[0, :, :]
            for i in range(dim[order]):
                var_Lambda = var3[:, :, i]
                inv_var_Lambda = inv((var_Lambda + var_Lambda.T)/2)
                vec = mvnrnd(np.matmul(inv_var_Lambda, var4[:, i]), inv_var_Lambda)
                if order == 0:
                    W[i, :] = vec.copy()
                elif order == 1:
                    X[i, :] = vec.copy()
                    
        if iters + 1 > maxiter1 - maxiter2:
            W_plus += W
            X_plus += X
            
        mat_hat = np.matmul(W, X.T)
        if iters + 1 > maxiter1 - maxiter2:
            mat_hat_plus += mat_hat
#         rmse = np.sqrt(np.sum((dense_mat[pos] - mat_hat[pos]) ** 2)/dense_mat[pos].shape[0])
        
        var_alpha = alpha + 0.5 * sparse_mat[position].shape[0]
        error = sparse_mat - mat_hat
        var_beta = beta + 0.5 * np.sum(error[position] ** 2)
        tau = np.random.gamma(var_alpha, 1/var_beta)

#         if (iters + 1) % 200 == 0 and iters < maxiter1 - maxiter2:
#             print('Iter: {}'.format(iters + 1))
#             print('RMSE: {:.6}'.format(rmse))
#             print()

    W = W_plus/maxiter2
    X = X_plus/maxiter2
    mat_hat = mat_hat_plus/maxiter2
#     if maxiter1 >= 100:
#         final_mape = np.sum(np.abs(dense_mat[pos] - mat_hat[pos])/dense_mat[pos])/dense_mat[pos].shape[0]
#         final_rmse = np.sqrt(np.sum((dense_mat[pos] - mat_hat[pos]) ** 2)/dense_mat[pos].shape[0])
#         print('Imputation MAPE: {:.6}'.format(final_mape))
#         print('Imputation RMSE: {:.6}'.format(final_rmse))
#         print()
    
    return mat_hat, W, X


def BTMF(dense_mat, sparse_mat, init, rank, time_lags, maxiter1, maxiter2):
    """Bayesian Temporal Matrix Factorization, BTMF."""
    W = init["W"]
    X = init["X"]
    
    d = time_lags.shape[0]
    dim1, dim2 = sparse_mat.shape
    pos = np.where((dense_mat != 0) & (sparse_mat == 0))
    position = np.where(sparse_mat != 0)
    binary_mat = np.zeros((dim1, dim2))
    binary_mat[position] = 1
    
    beta0 = 1
    nu0 = rank
    mu0 = np.zeros((rank))
    W0 = np.eye(rank)
    tau = 1
    alpha = 1e-6
    beta = 1e-6
    S0 = np.eye(rank)
    Psi0 = np.eye(rank * d)
    M0 = np.zeros((rank * d, rank))
    
    W_plus = np.zeros((dim1, rank))
    X_plus = np.zeros((dim2, rank))
    X_new_plus = np.zeros((dim2 + 1, rank))
    A_plus = np.zeros((rank, rank, d))
    mat_hat_plus = np.zeros((dim1, dim2 + 1))
    for iters in range(maxiter1):
        W_bar = np.mean(W, axis = 0)
        var_mu_hyper = (dim1 * W_bar)/(dim1 + beta0)
        var_W_hyper = inv(inv(W0) + cov_mat(W) + dim1 * beta0/(dim1 + beta0) * np.outer(W_bar, W_bar))
        var_Lambda_hyper = wishart(df = dim1 + nu0, scale = var_W_hyper, seed = None).rvs()
        var_mu_hyper = mvnrnd(var_mu_hyper, inv((dim1 + beta0) * var_Lambda_hyper))
        
        var1 = X.T
        var2 = kr_prod(var1, var1)
        var3 = tau * np.matmul(var2, binary_mat.T).reshape([rank, rank, dim1]) + np.dstack([var_Lambda_hyper] * dim1)
        var4 = (tau * np.matmul(var1, sparse_mat.T)
                + np.dstack([np.matmul(var_Lambda_hyper, var_mu_hyper)] * dim1)[0, :, :])
        for i in range(dim1):
            inv_var_Lambda = inv(var3[:, :, i])
            W[i, :] = mvnrnd(np.matmul(inv_var_Lambda, var4[:, i]), inv_var_Lambda)
        if iters + 1 > maxiter1 - maxiter2:
            W_plus += W
        
        Z_mat = X[np.max(time_lags) : dim2, :]
        Q_mat = np.zeros((dim2 - np.max(time_lags), rank * d))
        for t in range(np.max(time_lags), dim2):
            Q_mat[t - np.max(time_lags), :] = X[t - time_lags, :].reshape([rank * d])
        var_Psi = inv(inv(Psi0) + np.matmul(Q_mat.T, Q_mat))
        var_M = np.matmul(var_Psi, np.matmul(inv(Psi0), M0) + np.matmul(Q_mat.T, Z_mat))
        var_S = (S0 + np.matmul(Z_mat.T, Z_mat) + np.matmul(np.matmul(M0.T, inv(Psi0)), M0) 
                 - np.matmul(np.matmul(var_M.T, inv(var_Psi)), var_M))
        Sigma = invwishart(df = nu0 + dim2 - np.max(time_lags), scale = var_S, seed = None).rvs()
        A = mat2ten(mnrnd(var_M, var_Psi, Sigma).T, np.array([rank, rank, d]), 0)
        if iters + 1 > maxiter1 - maxiter2:
            A_plus += A

        Lambda_x = inv(Sigma)
        var1 = W.T
        var2 = kr_prod(var1, var1)
        var3 = tau * np.matmul(var2, binary_mat).reshape([rank, rank, dim2]) + np.dstack([Lambda_x] * dim2)
        var4 = tau * np.matmul(var1, sparse_mat)
        for t in range(dim2):
            Mt = np.zeros((rank, rank))
            Nt = np.zeros(rank)
            if t < np.max(time_lags):
                Qt = np.zeros(rank)
            else:
                Qt = np.matmul(Lambda_x, np.matmul(ten2mat(A, 0), X[t - time_lags, :].reshape([rank * d])))
            if t < dim2 - np.min(time_lags):
                if t >= np.max(time_lags) and t < dim2 - np.max(time_lags):
                    index = list(range(0, d))
                else:
                    index = list(np.where((t + time_lags >= np.max(time_lags)) & (t + time_lags < dim2)))[0]
                for k in index:
                    Ak = A[:, :, k]
                    Mt += np.matmul(np.matmul(Ak.T, Lambda_x), Ak)
                    A0 = A.copy()
                    A0[:, :, k] = 0
                    var5 = (X[t + time_lags[k], :] 
                            - np.matmul(ten2mat(A0, 0), X[t + time_lags[k] - time_lags, :].reshape([rank * d])))
                    Nt += np.matmul(np.matmul(Ak.T, Lambda_x), var5)
            var_mu = var4[:, t] + Nt + Qt
            if t < np.max(time_lags):
                inv_var_Lambda = inv(var3[:, :, t] + Mt - Lambda_x + np.eye(rank))
            else:
                inv_var_Lambda = inv(var3[:, :, t] + Mt)
            X[t, :] = mvnrnd(np.matmul(inv_var_Lambda, var_mu), inv_var_Lambda)
        mat_hat = np.matmul(W, X.T)
        
        X_new = np.zeros((dim2 + 1, rank))
        if iters + 1 > maxiter1 - maxiter2:
            X_new[0 : dim2, :] = X.copy()
            X_new[dim2, :] = np.matmul(ten2mat(A, 0), X_new[dim2 - time_lags, :].reshape([rank * d]))
            X_new_plus += X_new
            mat_hat_plus += np.matmul(W, X_new.T)
        
        tau = np.random.gamma(alpha + 0.5 * sparse_mat[position].shape[0], 
                              1/(beta + 0.5 * np.sum((sparse_mat - mat_hat)[position] ** 2)))
#         rmse = np.sqrt(np.sum((dense_mat[pos] - mat_hat[pos]) ** 2)/dense_mat[pos].shape[0])
#         if (iters + 1) % 200 == 0 and iters < maxiter1 - maxiter2:
#             print('Iter: {}'.format(iters + 1))
#             print('RMSE: {:.6}'.format(rmse))
#             print()

    W = W_plus/maxiter2
    X_new = X_new_plus/maxiter2
    A = A_plus/maxiter2
    mat_hat = mat_hat_plus/maxiter2
#     if maxiter1 >= 100:
#         final_mape = np.sum(np.abs(dense_mat[pos] - mat_hat[pos])/dense_mat[pos])/dense_mat[pos].shape[0]
#         final_rmse = np.sqrt(np.sum((dense_mat[pos] - mat_hat[pos]) ** 2)/dense_mat[pos].shape[0])
#         print('Imputation MAPE: {:.6}'.format(final_mape))
#         print('Imputation RMSE: {:.6}'.format(final_rmse))
#         print()
    
    return mat_hat, W, X_new, A


# -

def zeroFill(
    data,
    zero_fill_method='prevlatter',
    normalize=True,
    verbose=True,
):
    zero_count = np.sum(data == 0, axis=0)
    # Fill 0s in data
    if zero_fill_method == 'prevlatter':
        if verbose:
            print("{:^10}{:<30}:".format("", "Zero fill method"), "Previous then latter")
        filled_data = data.copy()
#         for j in range(filled_data.shape[1]):
#             if zero_count[j] == filled_data.shape[0]:
#                 filled_data[0, j] = 0.1
        for j in range(filled_data.shape[1]):
            for i in range(filled_data.shape[0]):
                if filled_data[i, j] == 0 and i >= 1:
                    filled_data[i, j] = filled_data[i - 1, j]
        for j in range(filled_data.shape[1] - 1, -1, -1):
            for i in range(filled_data.shape[0] - 1, -1, -1):
                if filled_data[i, j] == 0 and i <= filled_data.shape[0] - 2:
                    filled_data[i, j] = filled_data[i + 1, j]
#         nzc = np.sum(filled_data==0)
#         assert(nzc == 0)
#         print("jxr1")
        data = filled_data
    elif zero_fill_method in ['linear', 'nearest', 'zero', 'slinear', 'quadratic', 'cubic', 'previous', 'next']:
        # Possible interpolate methods are:
        # linear, nearest, zero, slinear, quadratic, cubic, previous, next
        if verbose:
            print(
                "{:^10}{:<30}:".format("", "Zero fill method"), zero_fill_method + " interpolate"
            )
        x = np.arange(data.shape[0])
        new_data = []
        for var in range(data.shape[1]):
            ind = data[:, var].nonzero()
            if ind[0].size > 1:
                f = interpolate.interp1d(
                    x[ind[0]],
                    data[ind[0], var],
                    kind=zero_fill_method,
                    fill_value="extrapolate"
                )
                new_data.append(f(x))
            elif ind[0].size == 1:
#                 import pdb; pdb.set_trace()
                new_data.append([data[ind[0], var].item()] * data.shape[0])
            else:
#                 import pdb; pdb.set_trace()
                new_data.append([0.0] * data.shape[0])
        data = np.array(new_data).T
    elif zero_fill_method == 'BPMF':
        X = data
        dim1, dim2 = X.shape
        rank = 10
        init = {"W": 0.1 * np.random.rand(dim1, rank), 
                "X": 0.1 * np.random.rand(dim2, rank)}
        maxiter1 = 100
        maxiter2 = 100
        fitMat, _W, _X = BPMF(X, X, init, rank, maxiter1, maxiter2)
        return fitMat
    elif zero_fill_method == 'BTMF':
        X = data
        dim1, dim2 = X.shape
        rank = min(dim1, dim2) // 2
        init = {"W": 0.1 * np.random.rand(dim1, rank), 
                "X": 0.1 * np.random.rand(dim2, rank)}
        maxiter1 = 1000
        maxiter2 = 100
        time_lags = np.array([1, 2, 3])
        fitMat, _W, _X, _A = BTMF(X, X, init, rank, time_lags, maxiter1, maxiter2)
        return fitMat
#     import pdb; pdb.set_trace()
    # Normalize data by subtract mean and divide by std
    if normalize:
        data_mean = np.mean(data, axis=0, keepdims=True)
        data_std = np.std(data, axis=0, keepdims=True)
        data_std[data_std == 0.0] = 1.0
        data = (data - data_mean) / data_std
    return data


def mask(
    data,
    sample_rate=0.8,
    mode='single', #single or section
    length=60,
    step=70,
    start=4353,
    segment=[0, 70, 140, 210, 280, 350, 420, 490, 560, 600]
):
    r = sample_rate
    if(mode == 'single'):
        dataMask = np.random.choice(2, data.shape, p=[1 - r, r])
        return data * dataMask, dataMask
    elif(mode == 'real'):
        oriR = (data != 0).sum() / np.size(data)
        nr = r / oriR
        loc = np.nonzero(data)
        nmask = np.random.choice(2, [loc[0].shape[0]], p=[1 - nr, nr])
        dataMask = np.ones(data.shape)
        dataMask[loc[0][nmask == 0], loc[1][nmask == 0]] = 0
        return data * dataMask, dataMask
    else:
        pass


def load(
    file_path,
    sheet_name="Sheet1",
    aggre_delta=1,
    normalize=True,
    zero_fill_method='prevlatter',
    verbose=True,
):
    """Load metric data from file_path. Each column is one variable.

    Params:
        file_path:
        sheet_name: name of sheet to load.
        normaliza: normalize data by subtract mean and divide by std.
        fill_zeros: fill 0 data with nearest sample.
        verbose: the debugging print level: 0 (Nothing), 1 (Method info), 2 (Phase info), 3(Algorithm info)

    Returns:
        data     : data in numpy array format, shape of [T, N], each column is a variable
        data_head: service names
    """
    # verbose >= 3, print data loading info
    if verbose and verbose >= 3:
        print("{:^10}{:<30}:".format("", "Data path"), file_path)
    # region Read excel sheet, each row of data is one variable
    wb = load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    data = []
    data_head = []
    for row_values in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True
    ):
        data_head.append(row_values[0])
        data.append(row_values[1:])

    if verbose and verbose >= 3:
        print("{:^10}{:<30}: ".format("", "Sheet Names"), end="")
        for name in wb.sheetnames:
            print(name, end=", ")
        print("")
        print("{:^10}{:<30}:".format("", "Loaded Sheet"), sheet_name)
        print(
            "{:^10}{:<30}:".format("", "Sheet Size"),
            "{} x {}".format(sheet.max_row, sheet.max_column),
        )
    wb.close()
    # endregion

    # region Aggregate data
    if aggre_delta != 1:
        # Aggregate data
        data = [aggregate(row, aggre_delta) for row in data]
    # transpose data, now each column is one variable
    data = np.array(data).T
    if verbose and verbose >= 3:
        print("{:^10}{:<30}:".format("", "Aggregate delta"), aggre_delta)
        print("{:^10}{:<30}:".format("", "Data Shape"), data.shape)
    # endregion

    zero_count = np.sum(data == 0, axis=0)
    # Fill 0s in data
    if zero_fill_method == 'prevlatter':
        if verbose:
            print("{:^10}{:<30}:".format("", "Zero fill method"), "Previous then latter")
        filled_data = data.copy()
        for j in range(filled_data.shape[1]):
            for i in range(filled_data.shape[0]):
                if filled_data[i, j] == 0 and i >= 1:
                    filled_data[i, j] = filled_data[i - 1, j]
        for j in range(filled_data.shape[1] - 1, -1, -1):
            for i in range(filled_data.shape[0] - 1, -1, -1):
                if filled_data[i, j] == 0 and i <= filled_data.shape[0] - 2:
                    filled_data[i, j] = filled_data[i + 1, j]
        data = filled_data
    elif zero_fill_method in ['linear', 'nearest', 'zero', 'slinear', 'quadratic', 'cubic', 'previous', 'next']:
        # Possible interpolate methods are:
        # linear, nearest, zero, slinear, quadratic, cubic, previous, next
        if verbose:
            print(
                "{:^10}{:<30}:".format("", "Zero fill method"), zero_fill_method + " interpolate"
            )
        x = np.arange(data.shape[0])
        new_data = []
        for var in range(data.shape[1]):
            ind = data[:, var].nonzero()
            f = interpolate.interp1d(
                x[ind[0]],
                data[ind[0], var],
                kind=zero_fill_method,
                fill_value="extrapolate",
            )
            new_data.append(f(x))
        data = np.array(new_data).T

    # Normalize data by subtract mean and divide by std
    if normalize:
        data_mean = np.mean(data, axis=0, keepdims=True)
        data_std = np.std(data, axis=0, keepdims=True)
        data = (data - data_mean) / data_std

    # print data attributes
    if verbose:
        print("{:^10}{:<30}:".format("", "Data header"))
        for i, head in enumerate(data_head):
            print("{:>15}({:4d} 0s):{}".format(i + 1, zero_count[i], head))
    return data, data_head
